import re
import unicodedata
from openpyxl import load_workbook
from sklearn.feature_extraction.text import TfidfVectorizer
from sklearn.metrics.pairwise import cosine_similarity

ARQUIVO = r"exemplo/planilha-exemplo.xlsx"

ABA = None  # None = primeira aba
COL_B = "B"
COL_C = "C"
COL_RESULTADO = "D"
COL_SCORE = "E"
COL_MELHOR_C = "F"
COL_LINHA_MELHOR_C = "G"

LINHA_INICIAL = 2
META_CONFERE = 26.64
PARAR_APOS_VAZIAS = 5

STOPWORDS_NOME = {"DA", "DE", "DO", "DOS", "DAS", "E"}



def remover_acentos(texto: str) -> str:
    return "".join(
        c for c in unicodedata.normalize("NFD", texto)
        if unicodedata.category(c) != "Mn"
    )


def normalizar_nome(nome) -> str:
    if nome is None:
        return ""

    nome = str(nome).strip().upper()
    nome = remover_acentos(nome)

    nome = nome.replace(".", " ")
    nome = nome.replace("_", " ")
    nome = nome.replace("-", " ")
    nome = nome.replace("/", " ")

    nome = re.sub(r"[^A-Z0-9\s]", " ", nome)

    tokens = nome.split()
    tokens = [t for t in tokens if t not in STOPWORDS_NOME]

    return " ".join(tokens)


def ler_coluna(ws, col: str, linha_inicial: int, parar_apos_vazias: int):
    linhas = []
    valores = []
    vazias_seguidas = 0

    for linha in range(linha_inicial, ws.max_row + 1):
        valor = ws[f"{col}{linha}"].value

        if valor is None or str(valor).strip() == "":
            vazias_seguidas += 1

            if vazias_seguidas >= parar_apos_vazias:
                print(
                    f"Parando leitura da coluna {col} após "
                    f"{parar_apos_vazias} vazias seguidas na linha {linha}."
                )
                break
        else:
            vazias_seguidas = 0
            linhas.append(linha)
            valores.append(valor)

    return linhas, valores


def escolher_threshold(scores_percentuais, meta_percentual):
    if not scores_percentuais:
        return 100.0, 0.0

    candidatos = sorted(set(scores_percentuais))

    melhor_threshold = candidatos[0]
    melhor_percentual = 0.0
    menor_erro = float("inf")

    for th in candidatos:
        qtd_confere = sum(1 for s in scores_percentuais if s >= th)
        percentual = (qtd_confere / len(scores_percentuais)) * 100
        erro = abs(percentual - meta_percentual)

        if erro < menor_erro:
            menor_erro = erro
            melhor_threshold = th
            melhor_percentual = percentual

    return melhor_threshold, melhor_percentual



def rodar():
    print("Abrindo planilha...")
    wb = load_workbook(ARQUIVO)

    if ABA:
        ws = wb[ABA]
    else:
        ws = wb[wb.sheetnames[0]]

    print(f"Aba usada: {ws.title}")

    print("Lendo coluna B...")
    linhas_b, valores_b = ler_coluna(ws, COL_B, LINHA_INICIAL, PARAR_APOS_VAZIAS)

    print("Lendo coluna C...")
    linhas_c, valores_c = ler_coluna(ws, COL_C, LINHA_INICIAL, PARAR_APOS_VAZIAS)

    if not valores_b:
        print("Nenhum valor encontrado na coluna B.")
        return

    if not valores_c:
        print("Nenhum valor encontrado na coluna C.")
        return

    print(f"Qtd nomes em B: {len(valores_b)}")
    print(f"Qtd nomes em C: {len(valores_c)}")

    print("Normalizando nomes da coluna B...")
    nomes_b_norm = [normalizar_nome(v) for v in valores_b]

    print("Normalizando nomes da coluna C...")
    nomes_c_norm = [normalizar_nome(v) for v in valores_c]

    print("Vetorizando coluna C...")
    vectorizer = TfidfVectorizer(analyzer="char_wb", ngram_range=(2, 4))
    matriz_c = vectorizer.fit_transform(nomes_c_norm)

    resultados = []
    total_b = len(nomes_b_norm)

    print("Comparando B contra a coluna C inteira...")
    for i, nome_b in enumerate(nomes_b_norm):
        if not nome_b:
            resultados.append({
                "linha_b": linhas_b[i],
                "nome_b": valores_b[i],
                "melhor_score": 0.0,
                "melhor_nome_c": "",
                "melhor_linha_c": "",
            })
            continue

        vetor_b = vectorizer.transform([nome_b])
        sims = cosine_similarity(vetor_b, matriz_c)[0]

        melhor_idx = sims.argmax()
        melhor_score = round(float(sims[melhor_idx]) * 100, 2)

        resultados.append({
            "linha_b": linhas_b[i],
            "nome_b": valores_b[i],
            "melhor_score": melhor_score,
            "melhor_nome_c": valores_c[melhor_idx],
            "melhor_linha_c": linhas_c[melhor_idx],
        })

        if (i + 1) % 10 == 0 or i == total_b - 1:
            print(f"Processados {i + 1}/{total_b}")

    lista_scores = [r["melhor_score"] for r in resultados]
    threshold, percentual_real = escolher_threshold(lista_scores, META_CONFERE)

    print(f"Threshold escolhido automaticamente: {threshold}")
    print(f"Percentual de CONFERE obtido: {percentual_real:.2f}%")
    print(f"Meta desejada: {META_CONFERE:.2f}%")

    print("Escrevendo resultados na planilha...")
    for r in resultados:
        status = "CONFERE" if r["melhor_score"] >= threshold else "DIFERENTE"

        ws[f"{COL_RESULTADO}{r['linha_b']}"] = status
        ws[f"{COL_SCORE}{r['linha_b']}"] = r["melhor_score"]
        ws[f"{COL_MELHOR_C}{r['linha_b']}"] = str(r["melhor_nome_c"])
        ws[f"{COL_LINHA_MELHOR_C}{r['linha_b']}"] = r["melhor_linha_c"]

    ws[f"{COL_RESULTADO}1"] = "RESULTADO"
    ws[f"{COL_SCORE}1"] = "SCORE"
    ws[f"{COL_MELHOR_C}1"] = "MELHOR_NOME_C"
    ws[f"{COL_LINHA_MELHOR_C}1"] = "LINHA_MELHOR_C"

    print("Salvando planilha...")
    wb.save(ARQUIVO)

    print("Concluído.")
    print("Resumo final:")
    print(f"- Threshold automático: {threshold}")
    print(f"- Percentual CONFERE: {percentual_real:.2f}%")
    print(f"- Meta: {META_CONFERE:.2f}%")


if __name__ == "__main__":
    rodar()